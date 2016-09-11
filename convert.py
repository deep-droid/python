import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook

# content is a string containing the file. For example the result of an http.request(url).
# You can also use a filepath by calling "xlrd.open_workbook(filepath)".
content = "./babolat/session171260.xls"
xlsBook = xlrd.open_workbook(file_contents=content)
workbook = openpyxlWorkbook()

for i in xrange(0, xlsBook.nsheets):
    xlsSheet = xlsBook.sheet_by_index(i)
    sheet = workbook.active if i == 0 else workbook.create_sheet()
    sheet.title = xlsSheet.name

    for row in xrange(0, xlsSheet.nrows):
        for col in xrange(0, xlsSheet.ncols):
            sheet.cell(row=row, column=col).value = xlsSheet.cell_value(row, col)