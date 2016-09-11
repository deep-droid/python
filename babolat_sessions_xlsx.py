import os
import xlrd
from openpyxl import load_workbook

count = 0
for file in os.listdir("./babolat/"):
	#if file.endswith("session171260.xls"):
	if file.endswith(".xlsx"):
		count+=1
		print(file)
		
		wb = load_workbook("./babolat/" + file)
		#print (wb.get_sheet_names())
		ws = wb.get_sheet_by_name('STROKES')
		#for row in worksheet2.iter_rows():
		#	print (str(row[0].value()))
		#for row in ws.iter_rows():
		#	for cell in row:
		#		print (cell.internal_value)
		
		# print(ws["b4"].value)
		# print(ws["b5"].value)
		# print(ws["b6"].value)
		# print(ws["b7"].value)

		# print(ws["c4"].value)
		# print(ws["c5"].value)
		# print(ws["c6"].value)
		# print(ws["c7"].value)
		
		# print(ws["d5"].value)
		# print(ws["d6"].value)

		# print(ws["b9"].value)
		# print(ws["b10"].value)
		# print(ws["b11"].value)
		# print(ws["b12"].value)

		# print(ws["c9"].value)
		# print(ws["c10"].value)
		# print(ws["c11"].value)
		# print(ws["c12"].value)
		
		# print(ws["d10"].value)
		# print(ws["d11"].value)
		
		#forehand
		FHall = ws["b4"].value
		FHtopspin = ws["b5"].value
		FHslice = ws["b6"].value
		FHflat = ws["b7"].value

		FHall_power = ws["c4"].value
		FHtopspin_power = ws["c5"].value
		FHslice_power = ws["c6"].value
		FHflat_power = ws["c7"].value

		FHtopspin_spin = ws["d5"].value
		FHslice_spin = ws["d5"].value

		#backhand
		BHall = ws["b9"].value
		BHtopspin = ws["b10"].value
		BHslice = ws["b11"].value
		BHflat = ws["b12"].value

		BHall_power = ws["c9"].value
		BHtopspin_power = ws["c10"].value
		BHslice_power = ws["c11"].value
		BHflat_power = ws["c12"].value

		BHtopspin_spin = ws["d10"].value
		BHslice_spin = ws["d11"].value	
		
		FHtopspin_percent = FHtopspin/(FHtopspin+FHflat)*100
		BHtopspin_percent = BHtopspin/(BHtopspin+BHflat)*100
		FHslice_percent = FHslice/(FHall)*100
		BHslice_percent = BHslice/(BHall)*100
		
		print(">> 	FH TS: %4.1f" % FHtopspin_percent + "% # " + "FH TS PW: %4.1f" % FHtopspin_power + " >> " +
		"BH TS: %4.1f" % BHtopspin_percent + "% # " + "BH TS PW: %4.1f" % BHtopspin_power +
		"FH SL: %4.1f" % FHslice_percent + "% # " + "BH SL: %4.1f" % BHslice_percent )
		#print(str("%f" % FHtopspin_percent).replace(".", ","))
		#print(str("%.1f" % round(FHtopspin_percent,1)))
print ("Sessions files: " + str(count))